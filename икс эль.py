from openpyxl import load_workbook


book = load_workbook('анчоус.xlsx')
sheet_1 = book['Лист1']
stikers_sheet = book['анчоус']

print(book.worksheets)
for i in range(1, 6):
    print(stikers_sheet.cell(row=i, column=1).value)


