from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook
TOKEN = '1604519780:AAF8-7AWybW-SOPVrBDjMBijv1aQr6Fu2j4'
book = load_workbook('еда.xlsx')
dish_sheet = book['Лист1']   #добавляю икс эль файл
stikers_sheet = book['Лист1']
fruhstuk_sheet = book['Завтраки и рыбные блюда']
flash_sheet = book['Мясные блюда']
vegetable_sheet = book['Вегетарианские блюда']

branches = ['Завтраки и рыбные блюда', 'Мясные блюда', 'Вегетарианские блюда']

dishes = {}
all_dishes = {}
for branch in branches:
    dishes[branch] = []
    i = 1
    while True:
        dish = book[branch].cell(row=i, column=1).value
        if dish is None:
            break
        dishes[branch].append(dish)
        all_dishes[dish] = (branch, i)
        i += 1
def main():
    u = Updater(token=TOKEN)  # обьект, который ловит сообщения в телеграм
    dispatcher = u.dispatcher
    handler = MessageHandler(Filters.all, do_echo)
    #help_handler = CommandHandler('help', do_help)
    start_handler = CommandHandler('start', do_start)
    keyboard_handler = MessageHandler(Filters.text, do_something)
    dispatcher.add_handler(start_handler)
    #dispatcher.add_handler(help_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(handler)
    u.start_polling()
# запускают бота (эти две строчки)
    u.idle()
def do_echo(update, context):
    update.message.reply_text(text='тут что-то наверно нужно') # создаем функцию повторения



def do_start(update, context):

    buttons = [[]]
    for branch in branches:
        if len(buttons[-1]) >= 3:
            buttons.append([branch])
        else:
            buttons[-1].append(branch)
    keyboard = buttons

    update.message.reply_text(
        text='Что угодно?',
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keydoard=True),
    )


def do_something(update, context):
    text = update.message.text.lower()
    i = 0
    for branch in branches:
        if text == branch.lower():
            buttons = [[]]
            for dish in dishes[branch]:
                if len(buttons[-1]) >= 3:
                    buttons.append([dish])
                else:
                    buttons[-1].append(dish)
            keyboard = buttons

            update.message.reply_text(
                text='Что именно?',
                reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keydoard=True),
            )
            return
    for dish in all_dishes.keys():
        if text == dish:
            branch, i = all_dishes[dish]
            components = book[branch].cell(row=i, column=2).value
            update.message.reply_text(components, reply_markup=ReplyKeyboardRemove())
            update.message.reply_text(
                text='Начать сначала?',
                reply_markup=ReplyKeyboardMarkup([['/start']], one_time_keyboard=True, resize_keydoard=True),
            )
            return
    update.message.reply_text('Без вариантов', reply_markup=ReplyKeyboardRemove())
    update.message.reply_text(
        text='Начать сначала?',
        reply_markup=ReplyKeyboardMarkup([['/start']], one_time_keyboard=True, resize_keydoard=True),
    )


if __name__ == '__main__':
    main()