from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook


TOKEN = '1604519780:AAF8-7AWybW-SOPVrBDjMBijv1aQr6Fu2j4'
book = load_workbook('анчоус.xlsx')
sheet_1 = book['Лист1']
stikers_sheet = book['анчоус']

def main():
    u = Updater(token=TOKEN)  # обьект, который ловит сообщения в телеграм

    dispatcher = u.dispatcher

    handler = MessageHandler(Filters.all, do_echo)
    help_handler = CommandHandler('help', do_help)
    start_handler = CommandHandler('start', do_start)
    keyboard_handler = MessageHandler(Filters.text, do_something)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(handler)
    u.start_polling() # запускают бота (эти две строчки)
    u.idle()


def do_echo(update, context):
    update.message.reply_text(text='ну привет')


def do_start(update, context):
    keyboard = [
        ['1', '2', '3'],
        ['привет', 'сыр', 'пока', 'майнкрафт'],
    ]
    update.message.reply_text(
        text=')))0))0)',
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keydoard=True),
     )


def do_something(update: Update, context):
    text = update.message.text

    print(stikers_sheet.max_row)
    for row in range(2, stikers_sheet.max_row + 1):
        catch_phrase = stikers_sheet.cell(row=row, column=4).value
        print(catch_phrase)
        print(text)

        if catch_phrase in text:

            stikers_id = stikers_sheet.cell(row=row, column=3).value
            update.message.reply_sticker(stikers_id)
            return


    if text == '1':
         update.message.reply_text('Вы нажали 1', reply_markup=ReplyKeyboardRemove())
    elif text == '2':
         update.message.reply_text('Вы нажали 2', reply_markup=ReplyKeyboardRemove())
    elif text == '3':
         update.message.reply_text('Вы нажали 3', reply_markup=ReplyKeyboardRemove())
    else:
         update.message.reply_text('Вы нажали что-то еще', reply_markup=ReplyKeyboardRemove())


def do_help(update: Update, context):
    user_id = update.message.from_user.id
    name = update.message.from_user.first_name
    update.message.reply_text(text=f'Привет, {name}!\n твой user_id: {user_id}.\n пока, пока')

def do_sticker(update: Update, context):
    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)
    update.message.reply_sticker(sticker_id)


main()

